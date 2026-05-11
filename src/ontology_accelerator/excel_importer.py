"""Import ontology definitions from Excel workbooks.

Expected Excel layout
---------------------
The workbook must contain an **Entities** sheet and may optionally contain a
**Relationships** sheet.  An optional **Ontology** sheet can set top-level
metadata.

**Ontology** sheet (optional):
    | Key          | Value                          |
    |--------------|--------------------------------|
    | displayName  | MyOntology                     |
    | description  | Created from Excel workbook    |
    | namespace    | usertypes                      |

**Entities** sheet – one row per property:
    | entity | property | type   | key | display | table    | schema |
    |--------|----------|--------|-----|---------|----------|--------|
    | Customer | customer_id | string | yes |       | customers | dbo   |
    | Customer | name        | string |     | yes   | customers | dbo   |

*   ``key`` – put ``yes`` / ``true`` / ``1`` to mark the key property.
*   ``display`` – put ``yes`` / ``true`` / ``1`` to mark the display property.
*   ``table`` / ``schema`` – optional; used to create data-bindings.

**Relationships** sheet (optional):
    | name                   | source_entity | target_entity | table  | sourceColumn | targetColumn |
    |------------------------|---------------|---------------|--------|--------------|--------------|
    | OrderBelongsToCustomer | Order         | Customer      | orders | order_id     | customer_id  |
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_TRUTHY = {"yes", "true", "1", "x"}


def _is_truthy(value) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() in _TRUTHY


def _cell_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _normalise_header(value: str) -> str:
    """Lower-case, strip, collapse whitespace, replace spaces with ``_``."""
    return "_".join(str(value).strip().lower().split())


def _read_sheet_as_dicts(ws) -> list[dict]:
    """Return each row of *ws* as a dict keyed by normalised column headers."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalise_header(h) for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        result.append({h: _cell_str(v) for h, v in zip(headers, row)})
    return result


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_excel(path: str | Path) -> dict:
    """Read an Excel workbook and return an ontology config dict.

    The returned dict has the same shape as a YAML-loaded config so it can be
    fed directly into ``validate_config`` and ``build_definition``.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    try:
        cfg: dict = {}

        # -- Ontology metadata --------------------------------------------------
        if "Ontology" in wb.sheetnames:
            meta_rows = _read_sheet_as_dicts(wb["Ontology"])
            meta = {}
            for row in meta_rows:
                key = row.get("key") or row.get("name") or row.get("field")
                val = row.get("value")
                if key and val:
                    meta[key.strip()] = val
            if meta:
                cfg["ontology"] = {
                    "displayName": meta.get("displayName", "OntologyFromExcel"),
                    "description": meta.get("description", "Created from Excel workbook"),
                    "namespace": meta.get("namespace", "usertypes"),
                }

        # -- Entities ------------------------------------------------------------
        if "Entities" not in wb.sheetnames:
            raise ValueError("Excel workbook must contain an 'Entities' sheet")

        entity_rows = _read_sheet_as_dicts(wb["Entities"])
        if not entity_rows:
            raise ValueError("The 'Entities' sheet is empty")

        # Group rows by entity name
        entities_map: dict[str, dict] = {}
        for row in entity_rows:
            entity_name = row.get("entity")
            prop_name = row.get("property")
            if not entity_name or not prop_name:
                continue

            if entity_name not in entities_map:
                entities_map[entity_name] = {
                    "name": entity_name,
                    "key": None,
                    "display": None,
                    "properties": [],
                    "_table": None,
                    "_schema": None,
                }

            entry = entities_map[entity_name]
            prop_type = row.get("type", "string") or "string"
            entry["properties"].append({"name": prop_name, "type": prop_type})

            if _is_truthy(row.get("key")):
                entry["key"] = prop_name
            if _is_truthy(row.get("display")):
                entry["display"] = prop_name

            # Binding metadata (same table for all props of this entity)
            if row.get("table") and not entry["_table"]:
                entry["_table"] = row["table"]
            if row.get("schema") and not entry["_schema"]:
                entry["_schema"] = row["schema"]

        entities: list[dict] = []
        for entry in entities_map.values():
            # Default key/display to first property if not set
            if not entry["key"] and entry["properties"]:
                entry["key"] = entry["properties"][0]["name"]
            if not entry["display"]:
                entry["display"] = entry["key"]

            entity: dict = {
                "name": entry["name"],
                "key": entry["key"],
                "display": entry["display"],
                "properties": entry["properties"],
            }

            # Build a data-binding if a table was specified
            table = entry.pop("_table", None)
            schema = entry.pop("_schema", None)
            if table:
                source: dict = {"sourceType": "LakehouseTable", "table": table}
                if schema:
                    source["schema"] = schema
                prop_bindings = {p["name"]: p["name"] for p in entry["properties"]}
                entity["bindings"] = [
                    {"source": source, "propertyBindings": prop_bindings}
                ]

            entities.append(entity)

        cfg["entities"] = entities

        # -- Relationships -------------------------------------------------------
        if "Relationships" in wb.sheetnames:
            rel_rows = _read_sheet_as_dicts(wb["Relationships"])
            relationships: list[dict] = []
            for row in rel_rows:
                name = row.get("name")
                src = row.get("source_entity")
                tgt = row.get("target_entity")
                if not name or not src or not tgt:
                    continue
                rel: dict = {
                    "name": name,
                    "source_entity": src,
                    "target_entity": tgt,
                }
                table = row.get("table")
                if table:
                    binding: dict = {
                        "source": {"sourceType": "LakehouseTable", "table": table},
                        "sourceColumn": row.get("source_column") or row.get("sourcecolumn"),
                        "targetColumn": row.get("target_column") or row.get("targetcolumn"),
                    }
                    if row.get("schema"):
                        binding["source"]["schema"] = row["schema"]
                    rel["bindings"] = [binding]
                relationships.append(rel)
            if relationships:
                cfg["relationships"] = relationships

        return cfg
    finally:
        wb.close()
